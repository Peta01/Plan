import tkinter as tk
import datetime 
import calendar
import locale
import pandas as pd
import holidays
from dateutil.easter import easter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os

locale.setlocale(locale.LC_ALL, 'cs_CZ')

def Bunka(col,row):
    return "{}{}".format(chr(col + 65),row+1)

col_letters = [ "A", "B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V"]
col_width =   [ 6, 4.5,  3,  3,  3,  3,  3,  3,  3,  5,  7,  5,  7,  6,  5,  7,  6,  5,  7,  6,  7,  6]

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Plánovací kalendář")
        arok = datetime.datetime.now().year
        zacatek = arok - 5
        konec = arok + 5
        self.years = []
        self.years = list(range(zacatek, konec + 1))
        self.selected_years = []
        
        self.create_checkboxes()
        self.create_buttons()
        self.rc = [65, 22]

        self.update()
        width = 267
        height = self.winfo_height()
        # print(height)
        # print(self.winfo_width())
        self.geometry("{}x{}".format(str(width),str(height)))

    def create_checkboxes(self):
        self.checkboxes_frame = tk.Frame(self)
        self.checkboxes_frame.pack(side=tk.LEFT)

        self.checkboxes = []
        for year in self.years:
            var = tk.IntVar()
            checkbox = tk.Checkbutton(self.checkboxes_frame, text=str(year), variable=var)
            checkbox.pack(anchor=tk.W)
            self.checkboxes.append((year, var))

    def create_buttons(self):
        buttons_frame = tk.Frame(self)
        buttons_frame.pack(side=tk.RIGHT, padx=10)

        ok_button = tk.Button(buttons_frame, text="OK", command=self.get_selected_years, width=10)
        ok_button.pack(pady=5, padx=5, anchor=tk.E)

        select_all_button = tk.Button(buttons_frame, text="Označit vše", command=self.select_all, width=10)
        select_all_button.pack(pady=5, padx=5, anchor=tk.E)

        deselect_all_button = tk.Button(buttons_frame, text="Zrušit vše", command=self.deselect_all, width=10)
        deselect_all_button.pack(pady=5, padx=5, anchor=tk.E)

        quit_button = tk.Button(buttons_frame, text="Konec", command=self.quit, width=10)
        quit_button.pack(pady=5, padx=5, anchor=tk.E)

    def get_selected_years(self):
#        self.configure(cursor="wait")
        self.selected_years = [year for year, var in self.checkboxes if var.get() == 1]
#        print(self.selected_years)
#        print(len(self.selected_years))
        if len(self.selected_years) == 0:
            return
        if len(self.selected_years) == 1:
            self.fname = "Plánovací kalendář {}.xlsx".format(str(self.selected_years[0]))
        else:
            self.fname = "Plánovací kalendář.xlsx"
#        print(self.fname)
        self.calendar = calendar.Calendar()
        # délka směny
        self.smena = 8
        wb = Workbook()
        wb.remove(wb.active)
        for year in self.selected_years:
            # přidání Velikonoc do svátků
            self.svatky = holidays.Czechia(years=year)
    #        print(self.svatky)
            pa = easter(year) + datetime.timedelta(days=-2)
            po = easter(year) + datetime.timedelta(days=1)
            self.svatky[pa] = "Velký pátek"
            self.svatky[po] = "Velikonoční pondělí"
            self.MakeCal(year)
            self.MakeWB(year, wb)
#        print(wb.sheetnames)
# Uložení sešitu
        wb.save(filename=self.fname)
#        self.configure(cursor="")
    # print(calendar.calendar(2023))
        os.startfile(self.fname)

    def select_all(self):
        for year, var in self.checkboxes:
            var.set(1)

    def deselect_all(self):
        for year, var in self.checkboxes:
            var.set(0)


    def MakeWB(self, year, wb):
        ws = wb.create_sheet(str(year))
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

         # configure the grid
        for r in range(0,self.rc[0]):
            for c in range(0,self.rc[1]):
                text = str(self.cal.loc[r, c])

                if text == "nan":
                    text = ""
                
                if self.span_cells.loc[r, c] != "ne":
                    if text.isnumeric():
                        ws[Bunka(c,r)] = int(text)
                    else:
                        ws[Bunka(c,r)] = text
                    cell = ws.cell(row=r+1, column=c+1)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                    cell.font = Font(bold=True)
                    if isinstance(self.span_cells.loc[r, c], list):
                        rs = self.span_cells.loc[r, c][0] - r + 1
                        cs = self.span_cells.loc[r, c][1] - c + 1
                        m_str = f"{Bunka(c,r)}:{Bunka(c+cs-1,r+rs-1)}"
                        ws.merge_cells(m_str)
                    den = self.days.loc[r,c]
                    if den == 1:
                        barva = "#ffff00"
                    elif den == 2:
                        barva = "#00ff00"
                    elif den == 3:
                        barva = "#ff0000"
                    else:
                        barva = "#f0f0f0"
                    if barva != "#f0f0f0":
                        barva = "ff"+barva[1:]
                        cell.fill = PatternFill(start_color=barva,
                               end_color=barva,
                               fill_type='solid')

        for i,col1 in enumerate(col_letters):
            ws.column_dimensions[col1].width = col_width[i]
            ws.page_setup.fitToPage = True
        ws.page_setup.horizontalCentered = True

    def MakeCal(self, year):
        tex = "{0}\n(+{1})"
        # self.cal obraz kalendáře v dataframe
        # self.span_cells dataframe sloučených buněk
        self.cal = pd.DataFrame(index=range(self.rc[0]),columns=range(self.rc[1]))
        self.span_cells = self.cal.copy()
        self.days = self.cal.copy()
        self.span_cells.loc[0,0] = [1,1]
        self.cal.loc[0,0] = year
        self.span_cells.loc[0,2] = [0,8]
        self.cal.loc[0,2] = "Dny v týdnu"
        self.span_cells.loc[0,9] = [0,10]
        self.cal.loc[0,9] = "Týden"
        self.span_cells.loc[0,11] = [0,13]
        self.cal.loc[0,11] = "Měsíc"
        self.span_cells.loc[0,14] = [0,16]
        self.cal.loc[0,14] = "Čtvrtletí"
        self.span_cells.loc[0,17] = [0,19]
        self.cal.loc[0,17] = "Pololetí"
        self.span_cells.loc[0,20] = [0,21]
        self.cal.loc[0,20] = "Rok"
        self.span_cells.loc[2,20] = [64,20]
        self.span_cells.loc[2,21] = [64,21]
        self.cal.loc[1,9] = "Číslo"
        self.cal.loc[1,10] = "Pr.dny"
        self.cal.loc[1,11] = "Číslo"
        self.cal.loc[1,12] = "Pr.dny"
        self.cal.loc[1,13] = "Hod."
        self.cal.loc[1,14] = "Číslo"
        self.cal.loc[1,15] = "Pr.dny"
        self.cal.loc[1,16] = "Hod."
        self.cal.loc[1,17] = "Číslo"
        self.cal.loc[1,18] = "Pr.dny"
        self.cal.loc[1,19] = "Hod."
        self.cal.loc[1,20] = "Pr.dny"
        self.cal.loc[1,21] = "Hod."
        
        # Dny v týdnu
        for i in range(7):
            self.cal.loc[1,i + 2] = calendar.day_abbr[i]

        # Naplnění dnů v měsících
        c = 2; r = 2
        cm = 0; cq = 14; cp = 17
        rq = 0; rp = 0
        workday_week = 0
        workday_month = 0
        workday_quarter = 0
        workday_half = 0
        workday_year = 0
        h_workday_week = 0
        h_workday_month = 0
        h_workday_quarter = 0
        h_workday_half = 0
        h_workday_year = 0
        for i in range(1,13):
            rm = r
            self.cal.loc[rm, cm] = calendar.month_name[i]
            self.cal.loc[rm, 11] = i
            if i == 1:
                self.cal.loc[r, cq] = 1
                self.cal.loc[r, cp] = 1
                rq = rp = r
            if i == 4:
                self.cal.loc[r, cq] = 2
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                rq = r
            if i == 7:
                self.cal.loc[r, cq] = 3
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                self.cal.loc[r, cp] = 2
                self.span_cells.loc[rp, cp] = [r - 1, cp]
                rq = rp = r
            if i == 10:
                self.cal.loc[r, cq] = 4
                rq = r
            for day in self.calendar.itermonthdays(year, i):
                if day != 0:
                    den = datetime.date(year, i, day)
                    dd = 0
                    if den.weekday() < 5:
                        dd = 0
                        if den in self.svatky:
                            dd = 3
                            h_workday_year += 1
                            h_workday_half += 1
                            h_workday_quarter +=1
                            h_workday_month += 1
                            h_workday_week += 1
                        else:
                            workday_year += 1
                            workday_half += 1
                            workday_quarter +=1
                            workday_month += 1
                            workday_week += 1

                    if den.weekday() == 5:
                        dd = 1
                    if den.weekday() == 6:
                        dd = 2
                    if den in self.svatky:
                        dd = 3
                    self.days.loc[r, c] = dd
                    self.cal.loc[r, c] = day
                    self.cal.loc[r, 9] = den.isocalendar().week
                    self.cal.loc[r, 10] = workday_week
                c += 1
                if c > 8:
                    workday_week = 0
                    h_workday_week = 0
                    c = 2
                    r += 1
            self.span_cells.loc[rm, cm] = [r - 1, 1]
            self.span_cells.loc[rm, 11] = [r - 1, 11]
            self.cal.loc[rm, 12] = tex.format(workday_month, h_workday_month)
            self.span_cells.loc[rm, 12] = [r - 1, 12]
            self.cal.loc[rm, 13] = tex.format(workday_month * self.smena, h_workday_month * self.smena)
            self.span_cells.loc[rm, 13] = [r - 1, 13]
            workday_month = 0
            h_workday_month = 0
            if i == 3:
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                self.cal.loc[rq, 15] = tex.format(workday_quarter, h_workday_quarter)
                self.span_cells.loc[rq, 15] = [r - 1, 15]
                self.cal.loc[rq, 16] = tex.format(workday_quarter * self.smena, h_workday_quarter * self.smena)
                self.span_cells.loc[rq, 16] = [r - 1, 16]
                workday_quarter = 0
                h_workday_quarter = 0
            if i == 6:
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                self.cal.loc[rq, 15] = tex.format(workday_quarter, h_workday_quarter)
                self.span_cells.loc[rq, 15] = [r - 1, 15]
                self.cal.loc[rq, 16] = tex.format(workday_quarter * self.smena, h_workday_quarter * self.smena)
                self.span_cells.loc[rq, 16] = [r - 1, 16]
                self.span_cells.loc[rp, cp] = [r - 1, cp]
                self.cal.loc[rp, 18] = tex.format(workday_half, h_workday_half)
                self.span_cells.loc[rp, 18] = [r - 1, 18]
                self.cal.loc[rp, 19] = tex.format(workday_half * self.smena, h_workday_half * self.smena)
                self.span_cells.loc[rp, 19] = [r - 1, 19]
                workday_quarter = 0
                h_workday_quarter = 0
                workday_half = 0
                h_workday_half = 0
            if i == 9:
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                self.cal.loc[rq, 15] = tex.format(workday_quarter, h_workday_quarter)
                self.span_cells.loc[rq, 15] = [r - 1, 15]
                self.cal.loc[rq, 16] = tex.format(workday_quarter * self.smena, h_workday_quarter * self.smena)
                self.span_cells.loc[rq, 16] = [r - 1, 16]
                workday_quarter = 0
                h_workday_quarter = 0
            if i == 12:
                self.span_cells.loc[rq, cq] = [r - 1, cq]
                self.cal.loc[rq, 15] = tex.format(workday_quarter, h_workday_quarter)
                self.span_cells.loc[rq, 15] = [r - 1, 15]
                self.cal.loc[rq, 16] = tex.format(workday_quarter * self.smena, h_workday_quarter * self.smena)
                self.span_cells.loc[rq, 16] = [r - 1, 16]
                self.span_cells.loc[rp, cp] = [r - 1, cp]
                self.cal.loc[rp, 18] = tex.format(workday_half, h_workday_half)
                self.span_cells.loc[rp, 18] = [r - 1, 18]
                self.cal.loc[rp, 19] = tex.format(workday_half * self.smena, h_workday_half * self.smena)
                self.span_cells.loc[rp, 19] = [r - 1, 19]
                self.cal.loc[2, 20] = tex.format(workday_year, h_workday_year)
                self.cal.loc[2, 21] = tex.format(workday_year * self.smena, h_workday_year * self.smena)
                workday_quarter = 0
                h_workday_quarter = 0
                workday_half = 0
                h_workday_half = 0
                workday_year = 0
                h_workday_year = 0
        
        for r in range(self.rc[0]):
            for c in range(self.rc[1]):
                if isinstance(self.span_cells.loc[r, c], list):
                    for r1 in range(r, self.span_cells.loc[r, c][0]+1):
                        for c1 in range(c, self.span_cells.loc[r, c][1]+1):
                            if not isinstance(self.span_cells.loc[r1, c1], list):
                                self.span_cells.loc[r1, c1] = "ne"


            # with pd.ExcelWriter("plan.xlsx") as writer:
            #     self.cal.to_excel(writer, sheet_name="plan")
            #     self.span_cells.to_excel(writer, sheet_name="span")
            #     self.days.to_excel(writer, sheet_name="days")


if __name__ == "__main__":
    app = App()
    app.mainloop()
